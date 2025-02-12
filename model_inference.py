# pmp_inter test
import torch
import argparse
import time
import os
import re
import math
import numpy as np
from torch.utils.data import DataLoader, TensorDataset, Dataset
from tqdm import tqdm
import torch.nn.functional as F
import gc
from einops import rearrange
import matplotlib.pyplot as plt
from tqdm import tqdm
from models.backbone import  MTT_mask_net, MTT_Net_HLG
import options.options as option
from utils.Map2Partition_qtmtt import map_to_parititon_qtmtt
import concurrent.futures
import matplotlib.pyplot as plt


raster2zscan4 = np.array([0, 1, 4, 5, 2, 3, 6, 7, 8, 9, 12, 13, 10, 11, 14, 15], dtype=np.int8)
work_on_PC = False

def str2bool(v):
    if isinstance(v, bool):
        return v
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')



def batch_index_select(x, idx):
    if len(x.size()) == 3:
        B, N, C = x.size()
        N_new = idx.size(1)
        offset = torch.arange(B, dtype=torch.long, device=x.device).view(B, 1) * N
        idx = idx + offset
        out = x.reshape(B * N, C)[idx.reshape(-1)].reshape(B, N_new, C)
        return out
    elif len(x.size()) == 2:
        B, N = x.size()
        N_new = idx.size(1)
        offset = torch.arange(B, dtype=torch.long, device=x.device).view(B, 1) * N
        idx = idx + offset
        out = x.reshape(B * N)[idx.reshape(-1)].reshape(B, N_new)
        return out
    else:
        raise NotImplementedError


def batch_index_fill(x, x1, x2, idx1, idx2):
    B, N, C = x.size()
    B, N1, C = x1.size()
    B, N2, C = x2.size()

    offset = torch.arange(B, dtype=torch.long, device=x.device).view(B, 1)
    idx1 = idx1 + offset * N
    idx2 = idx2 + offset * N

    x = x.reshape(B * N, C)

    x[idx1.reshape(-1)] = x1.reshape(B * N1, C)
    x[idx2.reshape(-1)] = x2.reshape(B * N2, C)

    x = x.reshape(B, N, C)
    return x


def load_ifo_from_cfg(cfg_path):
    fp = open(cfg_path)
    input_path = None
    bit_depth = None
    width = None
    height = None
    for line in fp:
        if "InputFile" in line:
            line = line.rstrip('\n').replace(' ', '').split('#')[0]
            loc = line.find(':')
            input_path = line[loc + 1:]
        elif "InputBitDepth" in line:
            bit_depth = int(line.rstrip('\n').replace(' ', '').split('#')[0].split(':')[-1])
        elif "SourceWidth" in line:
            width = int(line.rstrip('\n').replace(' ', '').split('#')[0].split(':')[-1])
        elif "SourceHeight" in line:
            height = int(line.rstrip('\n').replace(' ', '').split('#')[0].split(':')[-1])
        elif "FramesToBeEncoded" in line:
            frame_num = int(line.rstrip('\n').replace(' ', '').split('#')[0].split(':')[-1])
        elif "IntraPeriod" in line:
            intraperiod = int(line.rstrip('\n').replace(' ', '').split('#')[0].split(':')[-1])

    if (input_path is None) or (bit_depth is None) or (width is None) or (height is None) or (intraperiod is None):
        print("Format of CFG error !!!!!!!!")
        return
    return input_path, bit_depth, width, height, frame_num, intraperiod


def import_yuv420(file_path, width, height, frm_num, sub_sample_ratio=1, is10bit=False, gop_size=None):
    fp = open(file_path, 'rb')
    pix_num = width * height
    if gop_size:
        I_frm_num = frm_num // gop_size + 1
        sub_frm_num = frm_num - I_frm_num
    else:
        sub_frm_num = (frm_num + sub_sample_ratio - 1) // sub_sample_ratio  # actual frame number after downsampling
    if is10bit:
        data_type = np.uint16
    else:
        data_type = np.uint8
    y_temp = np.zeros(pix_num * sub_frm_num, dtype=data_type)
    u_temp = np.zeros(pix_num*sub_frm_num // 4, dtype=data_type)
    v_temp = np.zeros(pix_num*sub_frm_num // 4, dtype=data_type)

    for i in range(0, frm_num, sub_sample_ratio):
        if gop_size and i % gop_size == 0:
            continue
        if is10bit:
            fp.seek(i * pix_num * 3, 0)
        else:
            fp.seek(i * pix_num * 3 // 2, 0)
        if gop_size:
            subi = max(i - i // gop_size - 1, 0)
        else:
            subi = i // sub_sample_ratio
        y_temp[subi * pix_num: (subi + 1) * pix_num] = np.fromfile(fp, dtype=data_type, count=pix_num,
                                                                   sep='')  # u_temp[subi*pix_num//4 : (subi+1)*pix_num//4] = np.fromfile(fp, dtype=data_type, count=pix_num//4, sep='')  # v_temp[subi*pix_num//4 : (subi+1)*pix_num//4] = np.fromfile(fp, dtype=data_type, count=pix_num//4, sep='')

    fp.close()
    y = y_temp.reshape((sub_frm_num, height, width))
    u = u_temp.reshape((sub_frm_num, height//2, width//2))
    v = v_temp.reshape((sub_frm_num, height//2, width//2))
    # print(y.max(), u.max(), 'is_10bit ', is10bit)
    if is10bit:
        y = np.clip((y + 2) / 4, 0, 255).astype(np.uint8)
        u = np.clip((u + 2) / 4, 0, 255).astype(np.uint8)
        v = np.clip((v + 2) / 4, 0, 255).astype(np.uint8)
    if opt['datasets']['yuv']:
        return y, u, v  # return frm_num * H * W
    else:
        return y



def RA_recurrent(cur_id, a=0, b=32):
    if cur_id == a + (b - a) // 2:
        return a, b
    elif cur_id < a + (b - a) // 2:
        return RA_recurrent(cur_id, a=a, b=a + (b - a) // 2)
    elif cur_id > a + (b - a) // 2:
        return RA_recurrent(cur_id, a=a + (b - a) // 2, b=b)


def get_cand_id_list(cur_id, mode='RA', gop_size=32, ref_len=2, frm_num=32):
    """get candidate reference frame id list"""
    if mode == 'LDP' or mode == 'LDB':
        cand_id_list = set([max(0, (cur_id // gop_size - i) * gop_size) for i in range(4)])
        tmp_id = max(0, cur_id - 1)
        while len(cand_id_list) < 4:
            cand_id_list.add(tmp_id)
            tmp_id -= 1
            if tmp_id <= 0:
                break
        cand_id_list = list(cand_id_list)
        cand_id_list.sort(reverse=True)
        while len(cand_id_list) < 4:
            cand_id_list.append(cand_id_list[-1])
    elif mode == 'RA':
        base_id, remainder_id = cur_id // 32, cur_id % 32
        if base_id * 32 + 32 >= frm_num:
            f0_id, f1_id = RA_recurrent(cur_id=remainder_id, a=0, b=min(32, frm_num - base_id * 32))
        else:
            f0_id, f1_id = RA_recurrent(cur_id=remainder_id, a=0, b=32)
        cand_id_list = [base_id * 32 + f0_id, base_id * 32 + f0_id if (base_id * 32 + f1_id) >= frm_num else base_id * 32 + f1_id]  # 前向参考列表和后向参考列表的第一个frame id
    return cand_id_list[:ref_len]



def enc_dec_test(seq_name, qp, frm_num=64):
    cur_path = os.getcwd()
    # print(os.path.join(cur_path, "codec"))
    # exe_path = os.path.join(cur_path, 'codec')
    exe_path = os.path.join(cur_path)
    cfg_path = os.path.join(cur_path, "cfg")
    seq_path = os.path.join(opt['path']['test_seq_dir'], seq_name + '.yuv')

    profile_list = ['dpfast']
    for profile in profile_list:  # ["dpfast", "anchor"]:
        out_path = os.path.join(cur_path, "output", "RA", "QP" + str(qp), profile)
        log_path = os.path.join(cur_path, "log", "RA", "QP" + str(qp), profile)
        if not os.path.exists(log_path):
            os.makedirs(log_path)
        if not os.path.exists(out_path):
            os.makedirs(out_path)
        bin_name = seq_name + ".bin"
        enc_log_name = "enc_" + seq_name + ".log"
        dec_log_name = "dec_" + seq_name + ".log"
        if opt['enc_mode'] == 'LDP':
            encoder_type = '_LDP'
        elif opt['enc_mode'] == 'LDB':
            encoder_type = '_LDB'
        elif opt['enc_mode'] == 'RA':
            encoder_type = '_RA'
        cur_exe_path = os.path.join(exe_path, "TAppEncoder_" + profile + encoder_type + ".exe")
        mode_cfg_path = os.path.join(cfg_path, "encoder_randomaccess_vtm.cfg")  # RA mode
        cur_cfg_path = os.path.join(cfg_path, 'per-sequence', seq_name.split('_')[0] + ".cfg")
        cur_out_path = os.path.join(out_path, bin_name)
        cur_log_path = os.path.join(log_path, enc_log_name)

        enc_order = cur_exe_path + " -c " + mode_cfg_path + " -c " + cur_cfg_path + " -i " + seq_path + " -q " + str(qp) + " -b " + cur_out_path + " -f " + str(frm_num) + " > " + cur_log_path

        print(enc_order)
        # print(dec_order)
        os.system(enc_order)  # os.system(dec_order)


def depth2flag(depth_map, depth):
    cu_size = 64 >> depth  # 64 32 16
    cu_map_size = 8 >> depth  # 8 4 2
    offset = 1 << depth  # 1 2 4
    block_num = depth_map.shape[0]
    # print(block_num)
    flag_list = []
    for k in range(block_num):
        temp = []
        for i in range(0, 8, cu_map_size):
            for j in range(0, 8, cu_map_size):
                if depth_map[k, i, j] > depth:
                    temp.append(1)
                else:
                    temp.append(0)
        for i in range(len(temp)):
            if depth == 2:
                flag_list.append(temp[raster2zscan4[i]])
            else:
                flag_list.append(temp[i])
    return np.array(flag_list, dtype=np.int8)


def get_qt_flag(depth_map, save_name, qp):
    flag_map64 = depth2flag(depth_map, 0)
    flag_map32 = depth2flag(depth_map, 1)
    flag_map16 = depth2flag(depth_map, 2)
    save_name = os.path.join('./DepthFlag', save_name + '_Q' + str(qp))
    # print('save name:', save_name)
    out64_file = open(save_name + '_64.txt', 'w')
    out32_file = open(save_name + '_32.txt', 'w')
    out16_file = open(save_name + '_16.txt', 'w')
    for i in range(flag_map64.size):
        out64_file.write(str(flag_map64[i]) + '\n')
    for i in range(flag_map32.size):
        out32_file.write(str(flag_map32[i]) + '\n')
    for i in range(flag_map16.size):
        out16_file.write(str(flag_map16[i]) + '\n')
    out64_file.close()
    out32_file.close()
    out16_file.close()

def valid_qt_func(qt_map, depth):
    # input: 16x16
    if depth >= 4:
        return qt_map
    cu_size = 16 // (2 ** depth)
    map_size = qt_map.shape[-1]
    ratio_list = [0.2, 0.9, 0.9, 0.9]
    ratio = ratio_list[depth]
    depth_base = 1e-1  # 防止1e-6也变成
    for x in range(0, map_size, cu_size):
        for y in range(0, map_size, cu_size):
            if (qt_map[x:x+cu_size, y:y+cu_size] > depth+depth_base).sum() >= ratio * cu_size * cu_size:
                qt_map[x:x+cu_size, y:y+cu_size] = qt_map[x:x+cu_size, y:y+cu_size].clip(min=depth+1)
                qt_map[x:x+cu_size, y:y+cu_size] = valid_qt_func(qt_map[x:x+cu_size, y:y+cu_size], depth+1)
            else:
                qt_map[x:x+cu_size, y:y+cu_size] = qt_map[x:x+cu_size, y:y+cu_size].clip(max=depth)
    return qt_map


def get_sequence_partition_for_VTM(qt_map, bt_map, dire_map, is_luma, save_path, frm_num, frm_width, frm_height, qp, block_size=128, p_frame_id_list=None, label_qt=None, label_mtt=None, label_dire=None, ori_mtt_mask=None, m_label_depth_list=None, ):
    """qt_map shape(N, 3, 16, 16) / None bt_map shape(N, 4, 2, 32, 32)"""
    block_ratio = block_size // 64
    executor = concurrent.futures.ThreadPoolExecutor(max_workers=opt['thread_num'])
    future_results = []
    
    # out_file = open(save_path, 'w')
    block_num_in_height = frm_height // block_size
    block_num_in_width = frm_width // block_size
    seq_partition_hor_mat = np.zeros((frm_num, frm_height // 4, frm_width // 4))  # store whether is partition edge or not (1 or 0) for edges of all the basic unit (4*4)
    seq_partition_ver_mat = np.zeros((frm_num, frm_height // 4, frm_width // 4))
    seq_qt_map = np.zeros((frm_num, frm_height // 8, frm_width // 8))
    seq_dire_map = np.zeros((frm_num, 3, frm_height // 4, frm_width // 4))
    seq_mtt_mask = np.zeros((frm_num, frm_height // 128, frm_width // 128)).astype(np.int8)
    assert p_frame_id_list is not None
    for total_id, frm_id in tqdm(enumerate(p_frame_id_list)):
        # print("Frame ", frm_id)
        frm_block_id = total_id * block_num_in_height * block_num_in_width
        for block_x in range(block_num_in_height):
            for block_y in range(block_num_in_width):
                block_id = frm_block_id + block_x * block_num_in_width + block_y
                if ori_mtt_mask is not None:
                    seq_mtt_mask[frm_id, block_x, block_y] = int(ori_mtt_mask[block_id] * 100)          
                if bt_map is None:
                    bt_block = np.zeros((3,32,32))
                else:
                    bt_block = bt_map[block_id]
                if dire_map is None:
                    dire_block = np.zeros((3,32,32))
                else:
                    dire_block = dire_map[block_id]
                
                future = executor.submit(
                    map_to_parititon_qtmtt,
                    qt_map[block_id].clip(min=0, max=4), bt_block, dire_block, qp, chroma_factor=1, block_size=128, debug_mode=False,
                    no_dir=False, acc_level=opt['acc_level'], 
                    frm_id=frm_id, block_x=block_x, block_y=block_y
                )
                future_results.append(future)

    concurrent.futures.wait(future_results)
    results = []
    for future in future_results:
        result = future.result()
        results.append(result)
    for result in results:
        hor_mat, ver_mat, out_dire_map, valid_qt_map, frm_id, block_x, block_y = result
        seq_partition_hor_mat[frm_id, block_x * 16 * block_ratio:(block_x + 1) * 16 * block_ratio, block_y * 16 * block_ratio:(block_y + 1) * 16 * block_ratio] = hor_mat
        seq_partition_ver_mat[frm_id, block_x * 16 * block_ratio:(block_x + 1) * 16 * block_ratio, block_y * 16 * block_ratio:(block_y + 1) * 16 * block_ratio] = ver_mat
        seq_dire_map[frm_id, :, block_x * 16 * block_ratio:(block_x + 1) * 16 * block_ratio, block_y * 16 * block_ratio:(block_y + 1) * 16 * block_ratio] = out_dire_map
        seq_qt_map[frm_id, block_x * 8 * block_ratio:(block_x + 1) * 8 * block_ratio, block_y * 8 * block_ratio:(block_y + 1) * 8 * block_ratio] = valid_qt_map
    
    if save_path is not None:
        if os.path.exists(save_path):
            os.remove(save_path)
        out_file = open(save_path, 'w')
        # print(p_frame_id_list)
        for total_id, frm_id in tqdm(enumerate(p_frame_id_list)):
            hor_vec = seq_partition_hor_mat[frm_id].reshape(-1).astype(np.uint8)
            ver_vec = seq_partition_ver_mat[frm_id].reshape(-1).astype(np.uint8)
            qtdepth_vec = seq_qt_map[frm_id].reshape(-1).astype(np.uint8)
            dire_vec = seq_dire_map[frm_id].reshape(-1).astype(np.int8)
            if ori_mtt_mask is not None:
                mask_vec = seq_mtt_mask[frm_id].reshape(-1).astype(np.int8)
            for i in range(hor_vec.size):  # horizontal edge vector
                out_file.write(str(hor_vec[i]) + '\n')
            for i in range(ver_vec.size):  # vertical edge vector
                out_file.write(str(ver_vec[i]) + '\n')
            for i in range(qtdepth_vec.size):  # qt depth vector
                out_file.write(str(qtdepth_vec[i]) + '\n')
            for i in range(dire_vec.size):
                out_file.write(str(dire_vec[i]) + '\n')  # print(hor_vec.size)  # print(qtdepth_vec.size)  # print(dire_vec.size)
            if ori_mtt_mask is not None:
                for i in range(mask_vec.size):
                    out_file.write(str(mask_vec[i]) + '\n')
        out_file.close()  

    return seq_partition_hor_mat, seq_partition_ver_mat


def get_sequence_partition_for_VTM_debug(qt_map, bt_map, dire_map, is_luma, save_path, frm_num, frm_width, frm_height, qp, block_size=128, p_frame_id_list=None, label_qt=None, label_mtt=None, label_dire=None, ori_mtt_mask=None, m_label_depth_list=None, gop_size=None):
    """qt_map shape(N, 3, 16, 16) / None bt_map shape(N, 4, 2, 32, 32)"""
    valid_qt_map = None if qt_map is None else np.zeros_like(np.asarray(qt_map))
    valid_bt_map, valid_dire_map = np.zeros_like(bt_map), np.zeros_like(dire_map)
    block_num = bt_map.shape[0]
    qt_map = qt_map.clip(min=0, max=3)
    block_num_in_height = frm_height // block_size
    block_num_in_width = frm_width // block_size
    p_frame_id_list = list(set(range(frm_num)) - set(range(0, frm_num, gop_size)))
    seq_mtt_mask = np.zeros((frm_num, frm_height // 128, frm_width // 128)).astype(np.int8)
    for total_id, frm_id in tqdm(enumerate(p_frame_id_list)):
        # print("Frame ", frm_id)
        frm_block_id = total_id * block_num_in_height * block_num_in_width
        for block_x in range(block_num_in_height):
            for block_y in range(block_num_in_width):
                block_id = frm_block_id + block_x * block_num_in_width + block_y
                seq_mtt_mask[frm_id, block_x, block_y] = int(ori_mtt_mask[block_id] * 100)          
                if bt_map is None:
                    bt_block = np.zeros((3,32,32))
                else:
                    bt_block = bt_map[block_id]
                if dire_map is None:
                    dire_block = np.zeros((3,32,32))
                else:
                    dire_block = dire_map[block_id]
                
                valid_qt_block, valid_bt_block, valid_dire_block = map_to_parititon_qtmtt(qt_map[block_id].clip(min=0, max=4), bt_block, dire_block, qp, chroma_factor=1, block_size=128, debug_mode=True, no_dir=False, acc_level=opt['acc_level']) 
        
                valid_bt_map[block_id] = valid_bt_block
                valid_dire_map[block_id] = valid_dire_block
                valid_qt_map[block_id] = valid_qt_block
    # vis
    block_ratio = block_size // 64
    chroma_factor = 2
    if is_luma:
        chroma_factor = 1
    # assert save_path is not None
    block_num_in_height = frm_height // block_size
    block_num_in_width = frm_width // block_size

    inconsist_error = np.sum(valid_qt_map != qt_map.round()) / qt_map.size

    seq_bt_map = np.zeros((frm_num, 3, frm_height // 4, frm_width // 4))
    seq_qt_map = np.zeros((frm_num, frm_height // 8, frm_width // 8))
    seq_dire_map = np.zeros((frm_num, 3, frm_height // 4, frm_width // 4))
    # get p-frame id
    # p_frame_id_list = list(set(range(frm_num)) - set(range(0, frm_num, gop_size)))
    for total_id, frm_id in (enumerate(p_frame_id_list)):
        # print("Frame ", frm_id)
        frm_block_id = total_id * block_num_in_height * block_num_in_width
        for block_x in range(block_num_in_height):
            for block_y in range(block_num_in_width):
                block_id = frm_block_id + block_x * block_num_in_width + block_y

                seq_qt_map[frm_id, block_x * 8 * block_ratio:(block_x + 1) * 8 * block_ratio, block_y * 8 * block_ratio:(block_y + 1) * 8 * block_ratio] = valid_qt_map[block_id]
                seq_bt_map[frm_id, :, block_x * 16 * block_ratio:(block_x + 1) * 16 * block_ratio, block_y * 16 * block_ratio:(block_y + 1) * 16 * block_ratio] = valid_bt_map[block_id]
                seq_dire_map[frm_id, :, block_x * 16 * block_ratio:(block_x + 1) * 16 * block_ratio, block_y * 16 * block_ratio:(block_y + 1) * 16 * block_ratio] = valid_dire_map[block_id]
    
    
    print("AFTER post-processing")
    seq_qt_map = seq_qt_map[p_frame_id_list]
    seq_bt_map = seq_bt_map[p_frame_id_list]
    seq_dire_map = seq_dire_map[p_frame_id_list]
    accu_list, accu_mask = print_metric(m_label_depth_list=m_label_depth_list, out_pred_qt=seq_qt_map, out_pred_mtt=seq_bt_map, out_pred_dir=seq_dire_map, out_pred_mask=None)

    return seq_qt_map, seq_bt_map, seq_dire_map, accu_list, accu_mask, inconsist_error


def reblock_array(input_array, frame_num, frame_height, frame_width, origin_block_size=128, target_block_size=64, patch_size=16):
    """origin_block_size: origin luma block size, target_block_size: target luma block size, patch_size: map's patch size"""
    """input shape (N, C, 64, 64) output shape (N*4*4, C, 16, 16)  保持target_block的raster-scan"""
    block_ratio = origin_block_size // target_block_size
    target_array = np.zeros((input_array.shape[0] * block_ratio * block_ratio, input_array.shape[1], input_array.shape[2] // block_ratio, input_array.shape[3] // block_ratio))
    block_num_in_height, block_num_in_width = frame_height // origin_block_size, frame_width // origin_block_size
    for frm_id in range(frame_num):
        frm_block_id = frm_id * block_num_in_height * block_num_in_width
        for block_x in range(block_num_in_height):
            for block_y in range(block_num_in_width):
                block_id = frm_block_id + block_x * block_num_in_width + block_y
                target_parent_upper_id = (frm_block_id + block_x * block_num_in_width) * block_ratio * block_ratio
                for sub_height_id in range(block_ratio):
                    for sub_width_id in range(block_ratio):
                        target_block_id = target_parent_upper_id + sub_height_id * (frame_width // target_block_size) + block_y * block_ratio + sub_width_id  # 当前块上边的parent block行，当前块上边target_block行，当前块在当前行的index
                        target_array[target_block_id] = input_array[block_id, :, sub_height_id * patch_size: (sub_height_id + 1) * patch_size, sub_width_id * patch_size: (sub_width_id + 1) * patch_size]
    return target_array

def get_order(file_path):
    poc_list = []
    p_list = []
    reorder_list = []
    refer_id_f = {}
    refer_id_b = {}
    qp_dict = {}
    tid_dict = {}
    with open(file_path, 'r') as f:
        content = f.readlines()
    for line in content:
        if line[:3] == 'POC':
            pattern = r'POC\s+(\d+)'
            match = re.search(pattern, line)
            poc = int(match.group(1))
            poc_list.append(poc)
            if 'I-SLICE' not in line:
                p_list.append(poc)
            # qp
            qp_pattern = r"QP (\d+)"
            qp_match = re.search(qp_pattern, line)
            qp = int(qp_match.group(1))
            qp_dict[poc] = qp
            # tid
            tid_pattern = r'TId:\s*(\d+)'
            tid_match = re.search(tid_pattern, line)
            tid = int(tid_match.group(1))
            tid_dict[poc] = tid
            
            pattern = r'\[L0 ([\d\sc]+)\] \[L1 ([\d\sc]+)\]'
            matches = re.search(pattern, line)
            if matches:
                L0 = matches.group(1)
                L1 = matches.group(2)
                L0 = [int(item[:-1]) if item[-1] == "c" else int(item) for item in L0.split()]
                L1 = [int(item[:-1]) if item[-1] == "c" else int(item) for item in L1.split()]
                refer_id_f[poc] = L0
                refer_id_b[poc] = L1
            else:
                refer_id_b[poc] = 0
                refer_id_f[poc] = 0
    for i in range(len(poc_list)):
        reorder_list.append(poc_list.index(i))
    return refer_id_f, refer_id_b, reorder_list, qp_dict, tid_dict, sorted(p_list)


class MyDataset(Dataset):
    def __init__(self, file_path, width, height, frm_num, sub_sample_ratio, is10bit, mode, ref_len, enc_mode='RA', pad128=False, qp=None, opt=None, intra_period=None):
        super().__init__()
        assert intra_period is not None
        self.intra_period = intra_period
        self.enc_mode = enc_mode
        self.qp = qp
        self.gop_size = int(opt['encode_cfg'].split('gop')[-1])
        self.__load_dataset_y__( file_path, width, height, frm_num, sub_sample_ratio, is10bit, mode, ref_len, gop_size=self.gop_size, pad128=pad128)
        if opt['network']['qml'] or opt['network']['tml']:
            if work_on_PC:
                self.refer_id_f, self.refer_id_b, self.reorder_list, self.qp_dict, self.tid_dict, self.p_list = get_order(r"D:\VTM\VVCLAST\log\qp%d_frame64.log" % self.qp)
            else:
                self.refer_id_f, self.refer_id_b, self.reorder_list, self.qp_dict, self.tid_dict, self.p_list = get_order(f"/data/fengxm/pmp_dataset/log/{opt['encode_cfg']}/{opt['encode_cfg']}_ip{self.intra_period}_q{self.qp}.log")
        self.opt = opt

    def __load_dataset_y__(self,  file_path, width, height, frm_num, sub_sample_ratio, is10bit, mode, ref_len, gop_size, pad128):
        if opt['datasets']['yuv']:
            y,u,v = import_yuv420(file_path, width, height, frm_num, sub_sample_ratio, is10bit)
            u,v = u.repeat(2, axis=-1).repeat(2, axis=-2), v.repeat(2, axis=-1).repeat(2, axis=-2)
            self.content = np.concatenate([y[:,None],u[:,None],v[:,None]], axis=1)  
            if self.content.shape[-1]%128!=0 or self.content.shape[-2]%128!=0:
                if pad128:
                    content = np.zeros((self.content.shape[0], self.content.shape[1], math.ceil(self.content.shape[2] / 128) * 128,  math.ceil(self.content.shape[3] / 128) * 128))
                    content[:, :, :self.content.shape[2], :self.content.shape[3]] = self.content
                    self.content = content
                else:
                    h_cropped, w_cropped = self.content.shape[-2]//128*128, self.content.shape[-1]//128*128
                    self.content = self.content[:,:,:h_cropped, :w_cropped]
        else:
            y = import_yuv420(file_path, width, height, frm_num, sub_sample_ratio, is10bit)
            self.content = y[:,None]  
            if self.content.shape[-1]%128!=0 or self.content.shape[-2]%128!=0:
                if pad128:
                    content = np.zeros((self.content.shape[0], self.content.shape[1], math.ceil(self.content.shape[2] / 128) * 128,  math.ceil(self.content.shape[3] / 128) * 128))
                    content[:, :, :self.content.shape[2], :self.content.shape[3]] = self.content
                    self.content = content
                else:
                    h_cropped, w_cropped = self.content.shape[-2]//128*128, self.content.shape[-1]//128*128
                    self.content = self.content[:,:,:h_cropped, :w_cropped]

    def __getitem__(self, index):
        p_idx = index 
        abs_idx = self.p_list[p_idx]
        cand_frmid = [self.refer_id_f[abs_idx][0], self.refer_id_b[abs_idx][0]]
        i_frame = self.content[abs_idx]
        p0_frame = self.content[cand_frmid[0]]
        p1_frame = self.content[cand_frmid[1]]

        if self.opt['network']['qml'] or self.opt['network']['tml']:
            qp_list = [self.qp_dict[abs_idx], self.qp_dict[cand_frmid[0]], self.qp_dict[cand_frmid[1]]]
            tid_list = [self.tid_dict[abs_idx], self.tid_dict[cand_frmid[0]], self.tid_dict[cand_frmid[1]]]
            output = {'i_frame': i_frame, 'p0_frame': p0_frame, 'p1_frame':p1_frame, "qp_list": torch.tensor(qp_list), "tid_list": torch.tensor(tid_list)}
        else:
            output = {'i_frame': i_frame, 'p0_frame': p0_frame, 'p1_frame':p1_frame}
        return output  

    def __len__(self):
        return len(self.p_list)

    
def YUV2RGB(Y,U,V, isYUV420 = True):
    """Y: (frame_num, height, width), U/V: (frame_num, height//2, width//2)"""
    FRAME_NUM, IMG_HEIGHT, IMG_WIDTH = Y.shape[-3], Y.shape[-2], Y.shape[-1]
    bgr_data = torch.zeros(FRAME_NUM, 3, IMG_HEIGHT, IMG_WIDTH, dtype=torch.uint8).to(Y.device)
    if (isYUV420):
        U = F.interpolate(U, scale_factor=2)
        V = F.interpolate(V, scale_factor=2)

    c = (Y-16) * 298
    d = U - 128
    e = V - 128

    r = torch.floor((c + 409 * e + 128) / 256).long()
    g = torch.floor((c - 100 * d - 208 * e + 128)/256).long()
    b = torch.floor((c + 516 * d + 128)/256).long()

    r = torch.where(r < 0, 0, r)
    r = torch.where(r > 255,255,r)

    g = torch.where(g < 0, 0, g)
    g = torch.where(g > 255,255,g)

    b = torch.where(b < 0, 0, b)
    b = torch.where(b > 255,255,b)

    bgr_data[:, 2, :, :] = r
    bgr_data[:, 1, :, :] = g
    bgr_data[:, 0, :, :] = b
    # return (n,3,h,w)
    return bgr_data


@torch.no_grad()
def flow_norm(tensorFlow):
    Backward_tensorGrid_cpu = {}
    tensorHorizontal = torch.linspace(-1.0, 1.0, tensorFlow.size(3)).view(
        1, 1, 1, tensorFlow.size(3)).expand(tensorFlow.size(0), -1, tensorFlow.size(2), -1)
    tensorVertical = torch.linspace(-1.0, 1.0, tensorFlow.size(2)).view(
        1, 1, tensorFlow.size(2), 1).expand(tensorFlow.size(0), -1, -1, tensorFlow.size(3))
    Backward_tensorGrid_cpu[str(tensorFlow.size())] = torch.cat(
        [tensorHorizontal, tensorVertical], 1).cpu()

    tensorFlow = torch.cat([tensorFlow[:, 0:1, :, :] / ((tensorFlow.size(3) - 1.0) / 2.0),
                            tensorFlow[:, 1:2, :, :] / ((tensorFlow.size(2) - 1.0) / 2.0)], 1)
    return tensorFlow

@torch.no_grad()
def dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, label_batch_list=None, out_type='res', flag=None, p0_flow=None, p1_flow=None):
    """
    Returns:
    - input_batch: A batch containing the current frame's YUV texture, the residuals from warping the previous and current frames, 
      the residuals from warping the next and current frames, or the p_frame_aligned after 64x64 alignment.
    - flow: The optical flow and MDF from the previous frame, and the optical flow and MDF from the next frame.
    - When mtt-mask is set to true, the initial optical flow is reused from p0_flow and p1_flow.
    """
    if flag == 'mtt':
        Lm = 0
    elif flag == 'mtt_mask':
        Lm = 1
    elif flag == 'qt':
        Lm = 2
    
    if not opt['network']['light_spynet']:
        I_frame_YYY = F.interpolate(i_frame[:, 0:1], scale_factor=1 / ds, mode='nearest').repeat(1, 3, 1, 1) / 255.0
        P0_frame_YYY = F.interpolate(p0_frame[:, 0:1], scale_factor=1 / ds, mode='nearest').repeat(1, 3, 1, 1) / 255.0
        P1_frame_YYY = F.interpolate(p1_frame[:, 0:1], scale_factor=1 / ds, mode='nearest').repeat(1, 3, 1, 1) / 255.0
        P0_flow_list = flow_net(im1=I_frame_YYY, im2=P0_frame_YYY, Lm=Lm, init_flow=p0_flow)
        P1_flow_list = flow_net(im1=I_frame_YYY, im2=P1_frame_YYY, Lm=Lm, init_flow=p1_flow)
        if flag == 'qt':
            return P0_flow_list[-3], P1_flow_list[-3]  
        elif flag == 'mtt_mask':
            return [P0_flow_list[-2]], [P1_flow_list[-2]]
        elif flag == 'mtt':
            return [P0_flow_list[-3], P0_flow_list[-2], P0_flow_list[-1]], [P1_flow_list[-3], P1_flow_list[-2], P1_flow_list[-1]]
        else:
            raise Exception('invalid flag') 
    else:
        I_frame_YYY = F.interpolate(i_frame[:, 0:1], scale_factor=1 / ds, mode='nearest').repeat(1, 1, 1, 1) / 255.0
        P0_frame_YYY = F.interpolate(p0_frame[:, 0:1], scale_factor=1 / ds, mode='nearest').repeat(1, 1, 1, 1) / 255.0
        P1_frame_YYY = F.interpolate(p1_frame[:, 0:1], scale_factor=1 / ds, mode='nearest').repeat(1, 1, 1, 1) / 255.0
        p0_flow = flow_net(im1=I_frame_YYY, im2=P0_frame_YYY)
        p1_flow = flow_net(im1=I_frame_YYY, im2=P1_frame_YYY)
        return p0_flow, p1_flow
        

def metric_qt(out_qt, label_qt):
    accu, error = 0, 0
    accu = np.sum((out_qt.round() == label_qt)) / label_qt.size

    return accu, error


def print_metric(m_label_depth_list, out_pred_qt, out_pred_mtt=None, out_pred_dir=None, out_pred_mask=None):
    if out_pred_mask is not None:
        out_pred_mask = out_pred_mask.round()
    accu_list = [0, 0, 0, 0, 0, 0, 0]
    recall_list = [0, 0, 0, 0, 0, 0, 0]  
    zero_rate_list = [0,0,0,0,0,0, 0, 0] 
    zero_rate_list[0] = np.sum(out_pred_qt[:,::2,::2].round() == 0) / m_label_depth_list[0].size * 100
    recall_list[0] = np.sum((out_pred_qt[:,::2,::2].round() ==  m_label_depth_list[0].round()) * ( m_label_depth_list[0].round() != 0)) / np.sum( m_label_depth_list[0].round() != 0) * 100
    accu_list[0] = np.sum(out_pred_qt[:,::2,::2].round() ==  m_label_depth_list[0].round()) / out_pred_qt[:,::2,::2].size * 100
    
    accu_label = 0
    for pred_depth in range(3):
        mtt_depth = pred_depth + 1
        accu_label += m_label_depth_list[mtt_depth][:, 0]
        dir_label = m_label_depth_list[mtt_depth][:, 1]
        if out_pred_mtt is None:
            zero_rate_list[mtt_depth] = 0
            recall_list[mtt_depth] = 0
            accu_list[mtt_depth] = 0
        else:
            zero_rate_list[mtt_depth] = np.sum(out_pred_mtt[:, pred_depth].round() == 0) / accu_label.size * 100
            recall_list[mtt_depth] = np.sum((out_pred_mtt[:, pred_depth].round() ==  accu_label.round()) * ( out_pred_mtt[:, pred_depth].round() != 0)) / np.sum( out_pred_mtt[:, pred_depth].round() != 0) * 100
            accu_list[mtt_depth] = np.sum(out_pred_mtt[:, pred_depth].round() ==  accu_label.round()) / out_pred_mtt[:, pred_depth].size * 100

        if out_pred_dir is None:
            zero_rate_list[mtt_depth + 3] = 0
            recall_list[mtt_depth + 3] = 0
            accu_list[mtt_depth + 3] = 0
        else:
            zero_rate_list[mtt_depth + 3] = np.sum(out_pred_dir[:, pred_depth].round() == 0) / dir_label.size * 100
            recall_list[mtt_depth + 3] = np.sum((out_pred_dir[:, pred_depth].round() ==  dir_label.round()) * ( out_pred_dir[:, pred_depth].round() != 0)) / np.sum( out_pred_dir[:, pred_depth].round() != 0) * 100
            accu_list[mtt_depth + 3] = np.sum(out_pred_dir[:, pred_depth].round() ==  dir_label.round()) / out_pred_dir[:, pred_depth].size * 100

    # mt mask
    sum_mtt_weight = m_label_depth_list[1][:, 0] + m_label_depth_list[2][:, 0] + m_label_depth_list[3][:, 0]
    sum_mtt_weight = ((F.max_pool2d(torch.from_numpy(sum_mtt_weight).float(), kernel_size=4) + m_label_depth_list[0]) > torch.from_numpy(out_pred_qt[:,::2,::2]).round()) 

    mtt_mask_label = rearrange(rearrange(sum_mtt_weight, 'b (hi h) (wi w) -> b hi h wi w', h=8, w=8), 'b hi h wi w -> b hi wi h w').sum(dim=(3, 4)).float()
    mtt_mask_label = torch.clip(mtt_mask_label, min=0, max=1)
    
    if out_pred_mask is None:
        accu_mask = 0
        zero_mask = 0
    else:
        accu_mask = torch.sum(torch.from_numpy(out_pred_mask).flatten() == mtt_mask_label.flatten()) / mtt_mask_label.numel() * 100
        zero_mask = np.sum(out_pred_mask == 0) / mtt_mask_label.numel() * 100    
    
    print('TEST_METRIC of DEPTH ')
    for pred_depth in range(7):
        if pred_depth < 4:
            print("depth%d\t accuracy: " % pred_depth, '%.4f%%' % accu_list[pred_depth], "recall: ", \
                    '%.4f%%' % recall_list[pred_depth], "\t zero_rate: ", '%.4f%%' % zero_rate_list[pred_depth])
        else:
            print("dire%d\t accuracy: " % (pred_depth-4), '%.4f%%' % accu_list[pred_depth], "recall: ", \
                    '%.4f%%' % recall_list[pred_depth], "\t zero_rate: ", '%.4f%%' % zero_rate_list[pred_depth])
        if pred_depth == 3:
            print("--------------------------------")
    print("--------------------------------")
    print("mtt mask:\t accuracy: %.2f \t zero_rate: %.2f"%(accu_mask, zero_mask) )
            
    return accu_list, accu_mask



@torch.no_grad()
def inference_VVC(qp):
    print("QP:", qp)
    target_qp = qp
    if qp <= 24:
        target_qp = 22
    elif qp >= 25 and qp <= 29:
        target_qp = 27
    elif qp >= 30 and qp <= 34:
        target_qp = 32
    elif qp >= 35:
        target_qp = 37
    else:
        raise Exception('invalid qp')
    
    inference_time, post_time = {}, {}
    mask_ratio = {}
    print('stage: ', opt['stage'])
    qp_dict = {22:0, 27:1, 32:2, 37:3}
    qp_id = qp_dict[target_qp]
    stage = opt['stage']
    if target_qp == 22:
        text_fe = True
    else:
        text_fe = False
    if opt['network']['large_model'][qp_id]:
        if target_qp == 37:
            guide = False
        else:
            guide = True
        from models.backbone import QT_Net_HLG
        qt_net = QT_Net_HLG(spp=False, qml=True, tml=True, guide=guide)
    elif 'backbone' in opt['network'] and opt['network']['backbone'] != 'HourglassBlock':
        from models.backbone import QT_Net_HLG
        qt_net = QT_Net_HLG(qml=opt['network']['qml'], guide=True, backbone=opt['network']['backbone'])
    else:
        from models.backbone import QT_Net_HLG
        qt_net = QT_Net_HLG(qml=opt['network']['qml'], guide=True)
        
        
    from spynet.Spy_net import ME_Spynet
    flow_net = ME_Spynet(me_model_dir=opt['path']['me_model_dir'])

    mtt_mask_net = MTT_mask_net(qml=opt['network']['qml'], dlm=opt['network']['mtt_mask_dlm'])
    mtt_depth_net = MTT_Net_HLG(qml=opt['network']['qml'], residual_type=opt['network']['residual_type'], max_depth=3)
    mtt_dire_net = MTT_Net_HLG(qml=opt['network']['qml'], residual_type=opt['network']['residual_type'], max_depth=3)

    if torch.cuda.is_available() and opt['gpu_num'] > 0:
        print("Testing on GPU!")
        qt_net = qt_net.cuda()
        mtt_mask_net = mtt_mask_net.cuda()
        flow_net = flow_net.cuda()
        mtt_dire_net = mtt_dire_net.cuda()
        mtt_depth_net = mtt_depth_net.cuda()

    if opt['network']['large_model'][qp_id]:
        qt_net_path = os.path.join(opt['path']['qt_large_model_path'], 'model_qp%d.pth'%target_qp) if opt['path']['qt_model_path'][-4:] != ".pth" else opt['path']['qt_model_path']
    else:
        qt_net_path = os.path.join(opt['path']['qt_model_path'], 'model_qp%d.pth'%target_qp) if opt['path']['qt_model_path'][-4:] != ".pth" else opt['path']['qt_model_path']
        
    if stage >= 2:
        mtt_mask_net_path = os.path.join(opt['path']['mtt_mask_model_path'], 'model_qp%d.pth'%target_qp)
    
    if stage >= 3:
        mtt_depth_net_path = os.path.join(opt['path']['mtt_model_path'], 'model_qp%d.pth'%target_qp)
        mtt_dire_net_path = os.path.join(opt['path']['dire_model_path'], 'model_qp%d.pth'%target_qp)
        

    if torch.cuda.is_available() and opt['gpu_num'] > 0:
        if torch.cuda.device_count() > 1:
            qt_net.load_state_dict(torch.load(qt_net_path)['qt_net']) 
            if stage >= 2:
                mtt_mask_net.load_state_dict(torch.load(mtt_mask_net_path)['mtt_mask_net'])
            if stage >= 3:
                mtt_depth_net.load_state_dict(torch.load(mtt_depth_net_path)['mtt_mask_net'])
                mtt_dire_net.load_state_dict(torch.load(mtt_dire_net_path)['mtt_mask_net'])
        else:
            qt_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(qt_net_path)['qt_net'].items()})
            if stage >= 2:
                mtt_mask_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_mask_net_path)['mtt_mask_net'].items()})
            if stage >= 3:
                mtt_depth_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_depth_net_path)['mtt_depth_model'].items()})
                mtt_dire_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_dire_net_path)['mtt_dire_model'].items()})
    else:
        qt_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(qt_net_path, map_location=torch.device('cpu'))['qt_net'].items()})
        if stage >= 2:
            mtt_mask_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_mask_net_path, map_location=torch.device('cpu'))['mtt_mask_net'].items()})
        if stage >= 3:
            mtt_depth_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_depth_net_path, map_location=torch.device('cpu'))['mtt_depth_model'].items()})
            mtt_dire_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_dire_net_path, map_location=torch.device('cpu'))['mtt_dire_model'].items()})

    flow_net = flow_net.eval()
    qt_net = qt_net.eval()
    mtt_mask_net = mtt_mask_net.eval()
    mtt_depth_net = mtt_depth_net.eval()
    mtt_dire_net = mtt_dire_net.eval()

    if torch.cuda.is_available() and opt['gpu_num'] > 0:
        pass
    else:
        flow_net = flow_net.cpu()
        qt_net = qt_net.cpu()
        mtt_mask_net = mtt_mask_net.cpu()
        mtt_depth_net = mtt_depth_net.cpu()
        mtt_dire_net = mtt_dire_net.cpu()
    
    # load video
    seqs_info_path = "VVC_RA_Test_Sequences.txt"

    # seqs_info_fp = open(seqs_info_path, 'r')
    seqs = []
    with open(seqs_info_path, 'r') as seqs_info_fp:
        for line in seqs_info_fp:
            if line is None:
                break
            seqs.append(line.rstrip('\n').split(',')[0])

    # hyper parameters
    seq_accu_list = []
    for seq_id, seq_name in enumerate(seqs):
        if seq_id not in opt['datasets']['test_ids']:
            continue
        start_time = time.time()
        if 'RaceHorses' in seq_name and '832x480' in seq_name:
            cfg_path = '/data/fengxm/pmp_dataset/log/cfg/RaceHorsesC.cfg'
        else:
            cfg_path = os.path.join("/data/fengxm/pmp_dataset/log/cfg" if opt['path']['cfg_path'] is None else opt['path']['cfg_path'], seq_name.split('_')[0] + ".cfg")
        print("SEQ:", seq_name)
        seq_path, bit_depth, width, height, frame_num, intra_period = load_ifo_from_cfg(cfg_path)
        # continue
        # frame_num = opt['datasets']['frm_num']
        frame_num, gop_size = int(opt['encode_cfg'].split('_')[0].strip('f')), int(opt['encode_cfg'].split('_')[1].strip('gop'))
        if gop_size == 32 and intra_period == 48: 
            intra_period = 64
        width, height = int(seq_name.split('_')[1].split('x')[0]), int(seq_name.split('_')[1].split('x')[1])  # to make Horsers and HorsersC different
        seq_path = seq_name + '.yuv'
        if bit_depth != 10:
            is10bit = False
        else:
            is10bit = True
        input_path = os.path.join(opt['path']['test_seq_dir'], seq_path)
        if opt['boundary_handling']:
            width_round, height_round = math.ceil(width / 128) * 128, math.ceil(height / 128) * 128
            block_width_num, block_height_num = width_round // 128, height_round // 128
        else:
            block_width_num, block_height_num = width // 128, height // 128
            width_round, height_round = block_width_num * 128, block_height_num * 128
        if opt['module']['metric']:
            dataset_dir = opt['module']['metric_path']
            qt_label = np.load(os.path.join(dataset_dir, opt['enc_mode'] , 'QP%d'%target_qp, 'qt_label', '%s.npy'%seq_id))
            mtt_label = np.load(os.path.join(dataset_dir, opt['enc_mode'] , 'QP%d'%target_qp, 'mtt_label', '%s.npy'%seq_id))
            m_label_depth_list = [qt_label, mtt_label[:,0], mtt_label[:,1], mtt_label[:,2]]
        else:
            m_label_depth_list = None

        train_dataset = MyDataset(input_path, width, height, frm_num=frame_num, sub_sample_ratio=1, is10bit=is10bit, \
                                  ref_len=2 if opt['enc_mode'] == 'RA' else 8, mode=opt['enc_mode'], qp=qp, pad128=opt['boundary_handling'], opt=opt, intra_period=intra_period)
        train_data = DataLoader(train_dataset, batch_size=opt['datasets']['batchSize'], shuffle=False, num_workers=opt['datasets']['num_workers'], pin_memory=True, multiprocessing_context='spawn')
        p_frame_num = len(train_dataset.p_list)
        out_pred_qt = np.zeros((p_frame_num, height_round // 8, width_round // 8))
        out_pred_mt_mask = np.zeros((p_frame_num, height_round // 128, width_round // 128))
        out_pred_mtt = np.zeros((p_frame_num, 3, height_round // 4, width_round // 4))
        out_pred_dir = np.zeros((p_frame_num, 3, height_round // 4, width_round // 4))
        frm_id = 0

        for step, batch in tqdm(enumerate(train_data)):
            if torch.cuda.is_available() and opt['gpu_num'] > 0:
                i_frame, p0_frame, p1_frame = batch['i_frame'].cuda().float(), batch['p0_frame'].cuda().float(), batch['p1_frame'].cuda().float()  # input_batch shape(B, ref_frm_num+1, 384, 384), shape(B, ref_frm_num+1, 2, 192, 192)
                temp_qp_list = batch['qp_list'].cuda().float()
            else:
                i_frame, p0_frame, p1_frame = batch['i_frame'].float(), batch['p0_frame'].float(), batch['p1_frame'].float() 
                temp_qp_list = batch['qp_list'].float()
            if stage == 1:
                if opt['network']['large_model'][qp_id]: 
                    if opt['network']['light_spynet']:
                        p0_flow, p1_flow =  dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, flag='mtt')
                    else:
                        p0_flow, p1_flow =  dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, flag='mtt')
                        p0_flow, p1_flow = p0_flow[-1], p1_flow[-1]

                    qt_pred_list, _, _ = qt_net(i_frame[:, 0:1] / 255.0, torch.stack([p0_flow, p1_flow], dim=1), qp=temp_qp_list[:, 0:1],\
                        trans_flow_DAM=True, make_res=True, p0_frame=p0_frame[:, 0:1] / 255.0, p1_frame=p1_frame[:, 0:1] / 255.0, out_medium_feat=True, upsample=0)
                    
                else:
                    # new architecture
                    i_frame, p0_frame, p1_frame = i_frame[:, :, ::4, ::4], p0_frame[:, :, ::4, ::4], p1_frame[:, :, ::4, ::4]
                    p0_flow, p1_flow = dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, flag='qt')
                    qt_pred_list, qt_feature = qt_net(i_frame[:, 0:1] / 255.0, torch.stack([p0_flow, p1_flow], dim=1), qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=p0_frame[:, 0:1] / 255.0, p1_frame=p1_frame[:, 0:1] / 255.0, out_medium_feat=True, upsample=4)

                qt_pred_list = [ele * 4 - 0.5 for ele in qt_pred_list]
                qt_pred = qt_pred_list[-1]
                out_pred_qt[frm_id:frm_id+i_frame.shape[0]] = F.interpolate(qt_pred, scale_factor=2)[:,0].cpu()
                max_pred_depth = 1
            elif stage == 2:
                # mtt mask
                if opt['network']['large_model'][qp_id]: 
                    p0_flow, p1_flow =  dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, flag='mtt')
                    qt_pred_list, _, _ = qt_net(i_frame[:, 0:1] / 255.0, torch.stack([p0_flow[-1], p1_flow[-1]], dim=1), qp=temp_qp_list[:, 0:1],\
                        trans_flow_DAM=True, make_res=True, p0_frame=p0_frame[:, 0:1] / 255.0, p1_frame=p1_frame[:, 0:1] / 255.0, out_medium_feat=True, upsample=0)
                    p0_flow, p1_flow = p0_flow[-2], p1_flow[-2]
                else:
                    p0_flow, p1_flow = dataset2dataset(F.interpolate(i_frame, scale_factor=1/4, mode='bilinear'), F.interpolate(p0_frame, scale_factor=1/4, mode='bilinear'), F.interpolate(p1_frame, scale_factor=1/4, mode='bilinear'), flow_net, ds=1, flag='mtt')
                    qt_pred_list, _ = qt_net(F.interpolate(i_frame, scale_factor=1/4, mode='bilinear') / 255.0, torch.stack([p0_flow[-1], p1_flow[-1]], dim=1), qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=F.interpolate(p0_frame, scale_factor=1/4, mode='bilinear') / 255.0, p1_frame=F.interpolate(p1_frame, scale_factor=1/4, mode='bilinear') / 255.0, out_medium_feat=True, upsample=4)
                    p0_flow, p1_flow =  dataset2dataset(F.interpolate(i_frame, scale_factor=1/2, mode='bilinear'), F.interpolate(p0_frame, scale_factor=1/2, mode='bilinear'), F.interpolate(p1_frame, scale_factor=1/2, mode='bilinear'),\
                                                         flow_net, ds=1, flag='mtt_mask', p0_flow=p0_flow[-1], p1_flow=p1_flow[-1])
                    p0_flow, p1_flow = p0_flow[-1], p1_flow[-1]

                qt_pred_list = [ele * 4 - 0.5 for ele in qt_pred_list]
                qt_pred = qt_pred_list[-1]
                out_pred_qt[frm_id:frm_id+i_frame.shape[0]] = F.interpolate(qt_pred, scale_factor=2)[:,0].cpu()                   
                 
                mtt_mask_list = mtt_mask_net(F.interpolate(i_frame, scale_factor=1/2, mode='bilinear'), p0_flow=p0_flow, p1_flow=p1_flow, qt_pred=qt_pred, qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=F.interpolate(p0_frame, scale_factor=1/2, mode='bilinear'), p1_frame=F.interpolate(p1_frame, scale_factor=1/2, mode='bilinear'))
                mtt_mask_pred = mtt_mask_list[-1]
                out_pred_qt[frm_id:frm_id+i_frame.shape[0]] = F.interpolate(qt_pred, scale_factor=2)[:,0].cpu()
                out_pred_mt_mask[frm_id:frm_id+i_frame.shape[0]] = mtt_mask_pred.reshape(out_pred_mt_mask.shape[-2], out_pred_mt_mask.shape[-1]).cpu().numpy()
                
                max_pred_depth = 2
            elif stage >= 3:
                temp_qp_list = torch.ones_like(temp_qp_list) * target_qp
                if opt['network']['large_model'][qp_id]: 
                    p0_flow, p1_flow =  dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, flag='mtt')
                    qt_pred_list, _, _ = qt_net(i_frame[:, 0:1] / 255.0, torch.stack([p0_flow[-1], p1_flow[-1]], dim=1), qp=temp_qp_list[:, 0:1],\
                        trans_flow_DAM=True, make_res=True, p0_frame=p0_frame[:, 0:1] / 255.0, p1_frame=p1_frame[:, 0:1] / 255.0, out_medium_feat=True, upsample=0)
                else:
                    p0_flow, p1_flow = dataset2dataset(F.interpolate(i_frame, scale_factor=1/4, mode='bilinear'), F.interpolate(p0_frame, scale_factor=1/4, mode='bilinear'), F.interpolate(p1_frame, scale_factor=1/4, mode='bilinear'), flow_net, ds=1, flag='mtt')
                    qt_pred_list, _ = qt_net(F.interpolate(i_frame, scale_factor=1/4, mode='bilinear') / 255.0, torch.stack([p0_flow[-1], p1_flow[-1]], dim=1), qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=F.interpolate(p0_frame, scale_factor=1/4, mode='bilinear') / 255.0, p1_frame=F.interpolate(p1_frame, scale_factor=1/4, mode='bilinear') / 255.0, out_medium_feat=True, upsample=4)
                    p0_flow, p1_flow =  dataset2dataset(i_frame, p0_frame, p1_frame, flow_net, ds=1, flag='mtt', p0_flow=p0_flow[-1], p1_flow=p1_flow[-1])

                qt_pred_list = [ele * 4 - 0.5 for ele in qt_pred_list]
                qt_pred = qt_pred_list[-1]
                out_pred_qt[frm_id:frm_id+i_frame.shape[0]] = F.interpolate(qt_pred, scale_factor=2)[:,0].cpu()

                if target_qp == 27:
                    if seq_id == 10:
                        mtt_mask_net_path_tmp = os.path.join(opt['path']['mtt_mask_model_path'], 'model_qp22.pth')
                        if torch.cuda.is_available() and opt['gpu_num'] > 0:
                            if torch.cuda.device_count() > 1:
                                    mtt_mask_net.load_state_dict(torch.load(mtt_mask_net_path_tmp)['mtt_mask_net'])
                            else:
                                    mtt_mask_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_mask_net_path_tmp)['mtt_mask_net'].items()})
                        else:
                                mtt_mask_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_mask_net_path_tmp, map_location=torch.device('cpu'))['mtt_mask_net'].items()})
                    else:
                        mtt_mask_net_path_tmp = os.path.join(opt['path']['mtt_mask_model_path'], 'model_qp27.pth')
                        if torch.cuda.is_available() and opt['gpu_num'] > 0:
                            if torch.cuda.device_count() > 1:
                                    mtt_mask_net.load_state_dict(torch.load(mtt_mask_net_path_tmp)['mtt_mask_net'])
                            else:
                                    mtt_mask_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_mask_net_path_tmp)['mtt_mask_net'].items()})
                        else:
                                mtt_mask_net.load_state_dict({k.replace('module.', ''): v for k, v in torch.load(mtt_mask_net_path_tmp, map_location=torch.device('cpu'))['mtt_mask_net'].items()})

                mtt_mask_list = mtt_mask_net(F.interpolate(i_frame, scale_factor=1/2, mode='bilinear'), p0_flow=p0_flow[-2], p1_flow=p1_flow[-2], qt_pred=qt_pred, qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=F.interpolate(p0_frame, scale_factor=1/2, mode='bilinear'), p1_frame=F.interpolate(p1_frame, scale_factor=1/2, mode='bilinear'))
                mtt_mask_pred = mtt_mask_list[-1]
                out_pred_mt_mask[frm_id:frm_id+i_frame.shape[0]] = mtt_mask_pred.reshape(out_pred_mt_mask.shape[-2], out_pred_mt_mask.shape[-1]).cpu().numpy()
                
                mtt_depth_map_list, ctu_decision, drop_decision = mtt_depth_net(luma=i_frame[:, 0:1] / 255.0, p0_flow=p0_flow[-1], p1_flow=p1_flow[-1], qt_pred=qt_pred, qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=p0_frame / 255.0, p1_frame=p1_frame / 255.0, ctu_decision= mtt_mask_list[-1].view(i_frame.shape[0], -1), mask_ratio=0.3)
                mtt_dire_map_list, _, _ = mtt_dire_net(luma=i_frame[:, 0:1] / 255.0, p0_flow=p0_flow[-1], p1_flow=p1_flow[-1], qt_pred=qt_pred, qp=temp_qp_list[:, 0:1], trans_flow_DAM=True, p0_frame=p0_frame / 255.0, p1_frame=p1_frame / 255.0, ctu_decision= mtt_mask_list[-1].view(i_frame.shape[0], -1), mask_ratio=0.3)

                mtt_pred, mtt_dire_pred = [], []
                for layer_depth in range(3):

                    if opt['open_offset']:
                        offset = torch.zeros_like(mtt_depth_map_list[layer_depth])
                        offset[:,:,:,2] = opt['offset'][qp_id][layer_depth]
                        save_mtt_depth = torch.argmax(torch.softmax(mtt_depth_map_list[layer_depth], dim=-1) + offset, dim=-1)
                    else:
                        save_mtt_depth = torch.argmax(mtt_depth_map_list[layer_depth], dim=-1)
                        
                    mtt_pred_single = torch.zeros(i_frame.shape[0], i_frame.shape[-1] // 128 * i_frame.shape[-2] // 128, 32, 32)
                    drop_mtt_depth = torch.zeros(i_frame.shape[0], i_frame.shape[-1] // 128 * i_frame.shape[-2] // 128 - save_mtt_depth.shape[0], 32, 32)
                    if torch.cuda.is_available() and opt['gpu_num'] > 0:
                        mtt_pred_single, drop_mtt_depth = mtt_pred_single.cuda(), drop_mtt_depth.cuda()
                    mtt_pred_single = rearrange(mtt_pred_single, 'b n h w -> b n (h w)')
                    drop_mtt_depth = rearrange(drop_mtt_depth, 'b n h w -> b n (h w)')
                    save_mtt_depth = rearrange(save_mtt_depth, '(b n) h w -> b n (h w)', b=i_frame.shape[0]).float()
                    mtt_pred_single = batch_index_fill(mtt_pred_single, save_mtt_depth, drop_mtt_depth, ctu_decision, drop_decision)
                    mtt_pred.append(rearrange(mtt_pred_single, 'b (hi wi) (h w) -> b (hi h) (wi w)', hi=i_frame.shape[-2] // 128, wi=i_frame.shape[-1] // 128, h=32, w=32))
                    # dire
                    save_mtt_depth = torch.argmax(mtt_dire_map_list[layer_depth], dim=-1) - 1
                    mtt_pred_single = torch.zeros(i_frame.shape[0], i_frame.shape[-1] // 128 * i_frame.shape[-2] // 128, 32, 32)
                    drop_mtt_depth = torch.zeros(i_frame.shape[0], i_frame.shape[-1] // 128 * i_frame.shape[-2] // 128 - save_mtt_depth.shape[0], 32, 32)
                    if torch.cuda.is_available() and opt['gpu_num'] > 0:
                        save_mtt_depth, mtt_pred_single, drop_mtt_depth = save_mtt_depth.cuda(), mtt_pred_single.cuda(), drop_mtt_depth.cuda()
                    mtt_pred_single = rearrange(mtt_pred_single, 'b n h w -> b n (h w)')
                    drop_mtt_depth = rearrange(drop_mtt_depth, 'b n h w -> b n (h w)')
                    save_mtt_depth = rearrange(save_mtt_depth, '(b n) h w -> b n (h w)', b=i_frame.shape[0]).float()
                    mtt_pred_single = batch_index_fill(mtt_pred_single, save_mtt_depth, drop_mtt_depth, ctu_decision, drop_decision)
                    mtt_dire_pred.append(rearrange(mtt_pred_single, 'b (hi wi) (h w) -> b (hi h) (wi w)', hi=i_frame.shape[-2] // 128, wi=i_frame.shape[-1] // 128, h=32, w=32))
                    
                out_pred_mtt[frm_id:frm_id+i_frame.shape[0]] = torch.stack(mtt_pred, dim=1).cpu().numpy()
                mtt_dire_pred[0] = mtt_dire_pred[0] * (mtt_pred[0] > 0)
                out_pred_dir[frm_id:frm_id+i_frame.shape[0]] = torch.stack(mtt_dire_pred, dim=1).cpu().numpy()
                
            frm_id += i_frame.shape[0]
        if opt['module']['save_out_batch']:
            partition_batch_dir = os.path.join('/model/fengxm/pmp_model', 'PartitionBatch', opt['model_type'])
            if not os.path.exists(partition_batch_dir):
                os.makedirs(partition_batch_dir)

            np.save(os.path.join(partition_batch_dir, seq_name + '_Luma_QP' + str(qp) + '_qt.npy'), out_pred_qt)
            if stage >= 2:
                np.save(os.path.join(partition_batch_dir, seq_name + '_Luma_QP' + str(qp) + '_mtt_mask.npy'), out_pred_mt_mask)
            if stage >= 3:
                np.save(os.path.join(partition_batch_dir, seq_name + '_Luma_QP' + str(qp) + '_mtt.npy'), out_pred_mtt)
                np.save(os.path.join(partition_batch_dir, seq_name + '_Luma_QP' + str(qp) + '_dir.npy'), out_pred_dir)

        if opt['module']['metric']:
            print("SEQUENCE NAME: ", seq_name)
            print("BEFORE post-processing")
            accu_list, accu_mask = print_metric(m_label_depth_list=m_label_depth_list, out_pred_qt=out_pred_qt, out_pred_mtt=out_pred_mtt, out_pred_dir=out_pred_dir, out_pred_mask=out_pred_mt_mask)
            seq_accu_list.append(accu_list[0])
            pre_accu[qp].append([accu_list, accu_mask])
            
            
        model_infe_time = time.time() - start_time
        if opt['module']['infe_on_PC']:
            partition_mat_dir = os.path.join('/data/fengxm/pmp_dataset/', 'PartitionMat', opt['model_type'])
            if opt['path']['qt_candidate_path'] is not None:
                partition_mat_dir = os.path.join('/data/fengxm/pmp_dataset/', 'PartitionMat', opt['model_type'], qt_net_path.split('\\')[-1].split('.')[0])
            if not os.path.exists(partition_mat_dir):
                os.makedirs(partition_mat_dir)
            save_path = os.path.join(partition_mat_dir, seq_name + '_Luma_QP' + str(qp) + '_PartitionMat.txt')
            if opt['module']['recon_unconsist']:
                former_mask = np.zeros_like(out_total_batch_mtt[:, 4])
                for depth in [4, 3, 2, 1, 0]:
                    if depth == 0:
                        cur_mask = (out_total_batch_mtt[:, depth] != 0)
                    else:
                        cur_mask = ((out_total_batch_mtt[:, depth] - out_total_batch_mtt[:, depth - 1]) != 0)
                    correct_mask = ((former_mask == 1) * (cur_mask == 0))
                    for j in range(depth, 5):
                        out_total_batch_mtt[:, j][correct_mask] += 1
                    former_mask = ((out_total_batch_mtt[:, depth] - out_total_batch_mtt[:, depth - 1]) != 0)
                out_total_batch_mtt = np.clip(out_total_batch_mtt, a_min=0.0, a_max=7.0)
                del former_mask, cur_mask, correct_mask

            out_total_batch_qt = rearrange(out_pred_qt, 'f (hid h) (wid w) -> (f hid wid) h w', h=16, w=16) # .astype(np.uint8)
            if stage >= 1:
                out_total_batch_mask = rearrange(out_pred_mt_mask, 'f (hid h) (wid w) -> (f hid wid) h w', h=1, w=1) # .astype(np.uint8)
            if stage >= 2:
                out_total_batch_mtt = rearrange(out_pred_mtt, 'f c (hid h) (wid w) -> (f hid wid) c h w', h=32, w=32) # .astype(np.uint8)
                out_total_batch_dir = rearrange(out_pred_dir, 'f c (hid h) (wid w) -> (f hid wid) c h w', h=32, w=32) # .astype(np.uint8)
                out_total_batch_mtt[:,1] = out_total_batch_mtt[:,0] + out_total_batch_mtt[:,1]
                out_total_batch_mtt[:,2] = out_total_batch_mtt[:,2] + out_total_batch_mtt[:,1]
            
            if stage == 1:
                # # visualization
                # a, b = get_sequence_partition_for_VTM(out_total_batch_qt, None, None, is_luma=True, save_path=save_path, frm_num=frame_num, frm_width=width_round, frm_height=height_round, gop_size=32, qp=qp, label_qt=out_pred_qt, label_mtt=None, label_dire=None, ori_mtt_mask=out_total_batch_mask, m_label_depth_list=m_label_depth_list )
                # import matplotlib.pyplot as plt
                # plt.imsave('/code/qt_block_QP30.png', (selected_frame.detach().cpu().numpy()[0,0] / 255. * (1 - F.interpolate(torch.tensor(a+b).clamp(0,1).unsqueeze(1), mode='nearest', scale_factor=4)).numpy()[1,0] + (F.interpolate(torch.tensor(a+b).clamp(0,1).unsqueeze(1), mode='nearest', scale_factor=4)).numpy()[1,0]).clip(0,1)[128*3:128*6, 128*8:128*11], cmap='gray')
                get_sequence_partition_for_VTM(out_total_batch_qt, None, None, is_luma=True, save_path=save_path, frm_num=frame_num, frm_width=width_round, \
                                               frm_height=height_round, qp=target_qp, label_qt=out_pred_qt, label_mtt=None, label_dire=None, ori_mtt_mask=out_total_batch_mask, m_label_depth_list=m_label_depth_list, p_frame_id_list=train_dataset.p_list )
            else:
                if opt['module']['metric']:
                    seq_qt_map, seq_bt_map, seq_dire_map, accu_list, accu_mask, error = get_sequence_partition_for_VTM_debug(out_total_batch_qt, out_total_batch_mtt, \
                        out_total_batch_dir, is_luma=True, save_path=save_path, frm_num=frame_num, frm_width=width_round, frm_height=height_round, gop_size=gop_size, \
                            qp=target_qp, label_qt=out_pred_qt, label_mtt=out_pred_mtt, label_dire=out_pred_dir, ori_mtt_mask=out_total_batch_mask, m_label_depth_list=m_label_depth_list, p_frame_id_list=train_dataset.p_list)
                    post_accu[qp].append([accu_list, accu_mask])
                    inconsist_error[qp].append(error)
                else:
                    get_sequence_partition_for_VTM(out_total_batch_qt, out_total_batch_mtt, out_total_batch_dir, is_luma=True, save_path=save_path, frm_num=frame_num, \
                        frm_width=width_round, frm_height=height_round, qp=target_qp, label_qt=out_pred_qt, label_mtt=out_pred_mtt, label_dire=out_pred_dir, ori_mtt_mask=out_total_batch_mask, m_label_depth_list=m_label_depth_list, p_frame_id_list=train_dataset.p_list)
                
        infe_time = time.time() - start_time
        print("total inference time: %.4f"%infe_time, "model_inference_time: %.4f"%model_infe_time)  # enc_dec_test(seq_name, qp, frm_num=opt['datasets']['frm_num'])
        inference_time[seq_id] = model_infe_time
        if opt['show_mtt_mask']:
            mask_ratio[seq_id] = (out_pred_mt_mask < 0.5).sum() / len(out_pred_mt_mask.flatten())
        post_time[seq_id] = infe_time - model_infe_time
        gc.collect()
    if len(seq_accu_list) == 0:
        avg_accu = 0
    else:
        avg_accu = sum(seq_accu_list) / len(seq_accu_list)
    
    if opt['show_mtt_mask']:
        return inference_time, post_time, avg_accu, mask_ratio
    else:
        return inference_time, post_time, avg_accu


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-opt', type=str, help='Path to option YAML file.')
    args = parser.parse_args()
    opt = option.parse(args.opt, is_train=False)
    print(option.dict2str(opt))
    total_mtt_mask = []
    # inference
    qp_list = [22, 27, 32, 37]
    if opt['module']['metric']:
        pre_accu = {22: [], 27: [], 32: [], 37: []}
        post_accu = {22: [], 27: [], 32: [], 37: []}
        inconsist_error = {22: [], 27: [], 32: [], 37: []}
    total_infe_time, total_post_time = {}, {}
    if opt['show_mtt_mask']:
        total_mask_ratio = {}
        
    if opt['qp'] is None:
        for qp in qp_list:
            if opt['show_mtt_mask']:
                inference_time, post_time, avg_accu, mask_ratio = inference_VVC(qp)
            else:
                inference_time, post_time, avg_accu = inference_VVC(qp)
            total_infe_time[qp] = inference_time
            total_post_time[qp] = post_time
            total_mask_ratio[qp] = mask_ratio
        
        if opt['module']['metric']:
            record_accu, record_accu_post = {}, {}
            record_error = []
            for qp in [22,27,32,37]:
                seq_num, accu_list_sum, accu_mask_tmp = 0, 0, 0
                for accu_list, accu_mask in pre_accu[qp]:
                    accu_list_sum += np.asarray(accu_list)
                    accu_mask_tmp += accu_mask
                    seq_num += 1
                record_accu[qp] = np.concatenate([accu_list_sum / seq_num, np.asarray([accu_mask_tmp / seq_num])])
            for qp in [22,27,32,37]:
                seq_num, accu_list_sum, accu_mask_tmp = 0, 0, 0
                for accu_list, accu_mask in post_accu[qp]:
                    accu_list_sum += np.asarray(accu_list)
                    accu_mask_tmp += accu_mask
                    seq_num += 1
                record_accu_post[qp] = np.concatenate([accu_list_sum / seq_num, np.asarray([accu_mask_tmp / seq_num])])
            for qp in [22,27,32,37]:
                seq_num, accu_list_sum, accu_mask_tmp = 0, 0, 0
                record_error.append(np.mean(inconsist_error[qp]))
            print("QT results.")
            print(f"inconsist_error: {np.mean(record_error) * 100} %. pre_accu: {np.mean([record_accu[22][0], record_accu[27][0], record_accu[32][0], record_accu[37][0]])} %. post_accu: {np.mean([record_accu_post[22][0], record_accu_post[27][0], record_accu_post[32][0], record_accu_post[37][0]])} %.")
            for qp_id, qp in enumerate([22,27,32,37]):
                print(f"qp{qp}: error={record_error[qp_id] * 100} %. pre_accu: {record_accu[qp][0]} %. post_accu: {record_accu_post[qp][0]} %." )

        from openpyxl import load_workbook
        wb = load_workbook(filename='./results/inference.xlsx')
        ws = wb['Sheet1']
        # model_infer
        row_slice = 'C2:C89'  # model_inference CPU
        # row_slice = 'G2:G89'  # model_inference GPU bs=1
        # row_slice = 'H2:H89'  # model_inference GPU bs=4
        index = 0
        for row_id, row in enumerate(ws[row_slice]):
            for item_id,cell in enumerate(row):
                qp_id = row_id % 4
                seq_id = row_id // 4
                cell.value = total_infe_time[qp_list[row_id % 4]][row_id // 4]
            index += 1
        # post-algorithm
        row_slice = 'D2:D89'  # model_inference CPU
        index = 0
        for row_id, row in enumerate(ws[row_slice]):
            for item_id,cell in enumerate(row):
                qp_id = row_id % 4
                seq_id = row_id // 4
                cell.value = total_post_time[qp_list[row_id % 4]][row_id // 4]
            index += 1
        # # mask-ratio
        ws = wb['mask_ratio']
        row_slice = 'C2:C89'
        index = 0
        for row_id, row in enumerate(ws[row_slice]):
            for item_id,cell in enumerate(row):
                qp_id = row_id % 4
                seq_id = row_id // 4
                cell.value = total_mask_ratio[qp_list[row_id % 4]][row_id // 4]
            index += 1
        wb.save('./results/inference.xlsx') 
        wb.close()
    else:
        total_infe_time = inference_VVC(opt['qp'])

